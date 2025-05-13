import os
import sys
import sqlite3
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import filedialog, messagebox

# ========== SECCIÓN DE COLORES (EDITABLE) ==========
# Puedes cambiar los códigos ANSI aquí para modificar los colores
COLORS = {
    "menu": "\033[97m",         # Blanco
    "create": "\033[94m",       # Azul
    "xlsx2sqlite": "\033[93m",  # Amarillo
    "sqlite2xlsx": "\033[90m",  # Gris
    "error": "\033[91m",        # Rojo
    "success": "\033[95m",      # Morado
    "warning": "\033[38;5;208m",# Naranja (ANSI extendido)
    "progress": "\033[92m",     # Verde
    "reset": "\033[0m"
}

def colorize(msg, color_key):
    return f"{COLORS.get(color_key, COLORS['reset'])}{msg}{COLORS['reset']}"

# ========== GLOBAL: CONFIGURACIÓN DE TKINTER PARA SELECCIÓN DE ARCHIVOS ==========
def ask_save_file(title, defaultextension, filetypes):
    """Abre un diálogo para seleccionar ubicación y nombre de archivo para guardar."""
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.asksaveasfilename(
        title=title,
        defaultextension=defaultextension,
        filetypes=filetypes
    )
    root.destroy()
    return file_path

def ask_open_file(title, filetypes):
    """Abre un diálogo para seleccionar un archivo existente."""
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title=title,
        filetypes=filetypes
    )
    root.destroy()
    return file_path

def ask_directory(title):
    """Abre un diálogo para seleccionar una carpeta."""
    root = tk.Tk()
    root.withdraw()
    dir_path = filedialog.askdirectory(title=title)
    root.destroy()
    return dir_path

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

# ========== GLOBAL: LOG Y MANEJO DE SALIDA ==========
def print_log(msg, color="menu"):
    print(colorize(msg, color))

def wait_enter():
    input(colorize("\nPresione ENTER para continuar...", "menu"))

# ========== SECCIÓN 1: CREATE_MASTER_DB_XLSX ==========
def create_master_db_xlsx():
    """
    [CREATE_MASTER_DB_XLSX] Crea un archivo Excel maestro con estructura base.
    Pregunta al usuario dónde guardar el archivo (misma carpeta o seleccionar otra).
    """
    print_log("\n[CREATE_MASTER_DB_XLSX] Iniciando creación de archivo Excel maestro (db.xlsx)...", "create")
    print_log("Este proceso creará un archivo Excel con la estructura base para productos.", "create")
    print_log("¿Dónde desea guardar el archivo db.xlsx?", "create")
    print_log("1. En la misma carpeta donde se ejecuta este script.", "create")
    print_log("2. Seleccionar otra ubicación (se abrirá un navegador de archivos).", "create")
    while True:
        opcion = input(colorize("Seleccione una opción (1/2): ", "create")).strip()
        if opcion == "1":
            output_path = os.path.join(os.getcwd(), "db.xlsx")
            break
        elif opcion == "2":
            print_log("Seleccione la carpeta donde guardar el archivo db.xlsx...", "create")
            dir_path = ask_directory("Seleccione la carpeta para guardar db.xlsx")
            if dir_path:
                output_path = os.path.join(dir_path, "db.xlsx")
                break
            else:
                print_log("No se seleccionó ninguna carpeta. Intente de nuevo.", "warning")
        else:
            print_log("Opción inválida. Intente de nuevo.", "warning")

    # --- LÓGICA ORIGINAL DE create_master_db_xlsx.py ---
    wb = Workbook()
    ws = wb.active
    ws.append(['SKU', 'Título', 'EANs'])  # Encabezados
    ws.append(['123456', 'Producto de Ejemplo', '1234567890123,2345678901234'])  # Ejemplo
    try:
        wb.save(output_path)
        print_log(f"[CREATE_MASTER_DB_XLSX] Archivo creado exitosamente en: {output_path}", "success")
    except Exception as e:
        print_log(f"[CREATE_MASTER_DB_XLSX] ERROR al guardar el archivo: {e}", "error")
    wait_enter()

# ========== SECCIÓN 2: MIGRATE_DB_XLSX_TO_SQLITE ==========
def migrate_db_xlsx_to_sqlite():
    """
    [MIGRATE_DB_XLSX_TO_SQLITE] Migra datos desde un archivo db.xlsx a una base de datos SQLite db.db.
    Pregunta por la ubicación de ambos archivos de forma interactiva.
    """
    print_log("\n[MIGRATE_DB_XLSX_TO_SQLITE] Iniciando migración de Excel (db.xlsx) a SQLite (db.db)...", "xlsx2sqlite")
    # Selección de archivo db.xlsx
    print_log("¿Dónde se encuentra el archivo db.xlsx que desea migrar?", "xlsx2sqlite")
    print_log("1. En la misma carpeta donde se ejecuta este script.", "xlsx2sqlite")
    print_log("2. Seleccionar otra ubicación (se abrirá un navegador de archivos).", "xlsx2sqlite")
    while True:
        opcion = input(colorize("Seleccione una opción para db.xlsx (1/2): ", "xlsx2sqlite")).strip()
        if opcion == "1":
            xlsx_path = os.path.join(os.getcwd(), "db.xlsx")
            if not os.path.isfile(xlsx_path):
                print_log("No se encontró db.xlsx en la carpeta actual. Intente otra opción.", "warning")
                continue
            break
        elif opcion == "2":
            xlsx_path = ask_open_file("Seleccione el archivo db.xlsx", [("Archivos Excel", "*.xlsx")])
            if xlsx_path:
                break
            else:
                print_log("No se seleccionó ningún archivo. Intente de nuevo.", "warning")
        else:
            print_log("Opción inválida. Intente de nuevo.", "warning")

    # Selección de archivo db.db
    print_log("¿Dónde desea guardar la base de datos SQLite (db.db)?", "xlsx2sqlite")
    print_log("1. En la misma carpeta donde se ejecuta este script.", "xlsx2sqlite")
    print_log("2. Seleccionar otra ubicación (se abrirá un navegador de archivos).", "xlsx2sqlite")
    while True:
        opcion = input(colorize("Seleccione una opción para db.db (1/2): ", "xlsx2sqlite")).strip()
        if opcion == "1":
            db_path = os.path.join(os.getcwd(), "db.db")
            break
        elif opcion == "2":
            db_path = ask_save_file("Seleccione ubicación para db.db", ".db", [("Base de datos SQLite", "*.db")])
            if db_path:
                break
            else:
                print_log("No se seleccionó ninguna ubicación. Intente de nuevo.", "warning")
        else:
            print_log("Opción inválida. Intente de nuevo.", "warning")

    # --- LÓGICA ORIGINAL DE migrate_db_xlsx_to_sqlite.py ---
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS productos (
                sku TEXT PRIMARY KEY,
                titulo TEXT,
                eans TEXT
            )
        ''')
        wb = load_workbook(xlsx_path)
        ws = wb.active
        total_rows = ws.max_row - 1
        print_log(f'[MIGRATE_DB_XLSX_TO_SQLITE] Iniciando la migración de {total_rows} productos...', "progress")
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            sku, titulo, eans = row
            eans_combined = ','.join(e.strip() for e in str(eans).split(',')) if eans else ''
            cursor.execute('SELECT eans FROM productos WHERE sku = ?', (sku,))
            result = cursor.fetchone()
            if result:
                existing_eans = result[0]
                new_eans = ','.join(set(existing_eans.split(',') + eans_combined.split(',')))
                cursor.execute('UPDATE productos SET eans = ? WHERE sku = ?', (new_eans, sku))
            else:
                cursor.execute('INSERT OR IGNORE INTO productos (sku, titulo, eans) VALUES (?, ?, ?)', (sku, titulo, eans_combined))
            print_log(f'Progreso: {idx}/{total_rows} productos migrados', "progress")
        conn.commit()
        conn.close()
        print_log('[MIGRATE_DB_XLSX_TO_SQLITE] Migración completada.', "success")
    except Exception as e:
        print_log(f'[MIGRATE_DB_XLSX_TO_SQLITE] ERROR: {e}', "error")
    wait_enter()

# ========== SECCIÓN 3: MIGRATE_DB_SQLITE_TO_XLSX ==========
def migrate_db_sqlite_to_xlsx():
    """
    [MIGRATE_DB_SQLITE_TO_XLSX] Migra datos desde una base de datos SQLite db.db a un archivo Excel db.xlsx.
    Pregunta por la ubicación de ambos archivos de forma interactiva.
    """
    print_log("\n[MIGRATE_DB_SQLITE_TO_XLSX] Iniciando migración de SQLite (db.db) a Excel (db.xlsx)...", "sqlite2xlsx")
    # Selección de archivo db.db
    print_log("¿Dónde se encuentra la base de datos SQLite (db.db) que desea migrar?", "sqlite2xlsx")
    print_log("1. En la misma carpeta donde se ejecuta este script.", "sqlite2xlsx")
    print_log("2. Seleccionar otra ubicación (se abrirá un navegador de archivos).", "sqlite2xlsx")
    while True:
        opcion = input(colorize("Seleccione una opción para db.db (1/2): ", "sqlite2xlsx")).strip()
        if opcion == "1":
            db_path = os.path.join(os.getcwd(), "db.db")
            if not os.path.isfile(db_path):
                print_log("No se encontró db.db en la carpeta actual. Intente otra opción.", "warning")
                continue
            break
        elif opcion == "2":
            db_path = ask_open_file("Seleccione el archivo db.db", [("Base de datos SQLite", "*.db")])
            if db_path:
                break
            else:
                print_log("No se seleccionó ningún archivo. Intente de nuevo.", "warning")
        else:
            print_log("Opción inválida. Intente de nuevo.", "warning")

    # Selección de archivo db.xlsx de salida
    print_log("¿Dónde desea guardar el archivo Excel (db.xlsx) de salida?", "sqlite2xlsx")
    print_log("1. En la misma carpeta donde se ejecuta este script.", "sqlite2xlsx")
    print_log("2. Seleccionar otra ubicación (se abrirá un navegador de archivos).", "sqlite2xlsx")
    while True:
        opcion = input(colorize("Seleccione una opción para db.xlsx (1/2): ", "sqlite2xlsx")).strip()
        if opcion == "1":
            xlsx_path = os.path.join(os.getcwd(), "db.xlsx")
            break
        elif opcion == "2":
            xlsx_path = ask_save_file("Seleccione ubicación para db.xlsx", ".xlsx", [("Archivos Excel", "*.xlsx")])
            if xlsx_path:
                break
            else:
                print_log("No se seleccionó ninguna ubicación. Intente de nuevo.", "warning")
        else:
            print_log("Opción inválida. Intente de nuevo.", "warning")

    # --- LÓGICA ORIGINAL DE migrate_db_sqlite_to_xlsx.py ---
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        wb = Workbook()
        ws = wb.active
        ws.append(['SKU', 'Titulo', 'EANs'])
        cursor.execute('SELECT sku, titulo, eans FROM productos')
        productos = cursor.fetchall()
        total_productos = len(productos)
        print_log(f'[MIGRATE_DB_SQLITE_TO_XLSX] Iniciando la migración de {total_productos} productos...', "progress")
        for idx, producto in enumerate(productos, start=1):
            ws.append(producto)
            print_log(f'Progreso: {idx}/{total_productos} productos migrados', "progress")
        wb.save(xlsx_path)
        conn.close()
        print_log('[MIGRATE_DB_SQLITE_TO_XLSX] Migración completada de SQLite a Excel.', "success")
    except Exception as e:
        print_log(f'[MIGRATE_DB_SQLITE_TO_XLSX] ERROR: {e}', "error")
    wait_enter()

# ========== GLOBAL: MENÚ PRINCIPAL ==========
def main_menu():
    while True:
        clear_screen()
        print_log("===============================================", "menu")
        print_log("  HERRAMIENTA INTERACTIVA DE MIGRACIÓN DB/EXCEL", "menu")
        print_log("===============================================", "menu")
        print_log("Seleccione una opción:", "menu")
        print_log("1. Crear archivo maestro Excel (create_master_db_xlsx)", "menu")
        print_log("2. Migrar de Excel a SQLite (migrate_db_xlsx_to_sqlite)", "menu")
        print_log("3. Migrar de SQLite a Excel (migrate_db_sqlite_to_xlsx)", "menu")
        print_log("4. Salir", "menu")
        print_log("-----------------------------------------------", "menu")
        opcion = input(colorize("Ingrese el número de la opción deseada: ", "menu")).strip()
        if opcion == "1":
            print_log("\n[Menú] Opción seleccionada: Crear archivo maestro Excel", "create")
            create_master_db_xlsx()
        elif opcion == "2":
            print_log("\n[Menú] Opción seleccionada: Migrar de Excel a SQLite", "xlsx2sqlite")
            migrate_db_xlsx_to_sqlite()
        elif opcion == "3":
            print_log("\n[Menú] Opción seleccionada: Migrar de SQLite a Excel", "sqlite2xlsx")
            migrate_db_sqlite_to_xlsx()
        elif opcion == "4":
            print_log("\nSaliendo de la aplicación. ¡Hasta luego!", "success")
            break
        else:
            print_log("Opción inválida. Intente de nuevo.", "warning")
            wait_enter()

# ========== GLOBAL: MANEJO DE CTRL+C ==========
def main():
    try:
        main_menu()
    except KeyboardInterrupt:
        print_log("\n\n[GLOBAL] Interrupción detectada (Ctrl+C). Saliendo de la aplicación.", "error")
        sys.exit(0)

if __name__ == "__main__":
    main()
