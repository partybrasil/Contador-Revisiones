import sqlite3
from openpyxl import load_workbook

# Conectar a la base de datos SQLite
conn = sqlite3.connect('db.db')
cursor = conn.cursor()

# Crear la tabla si no existe
cursor.execute('''
    CREATE TABLE IF NOT EXISTS productos (
        sku TEXT PRIMARY KEY,
        titulo TEXT,
        eans TEXT
    )
''')

# Cargar datos desde db.xlsx
wb = load_workbook('db.xlsx')
ws = wb.active

# Contar el número total de filas para mostrar el progreso
total_rows = ws.max_row - 1
print(f'Iniciando la migración de {total_rows} productos...')

# Insertar datos en la base de datos SQLite
for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
    sku, titulo, eans = row
    eans_combined = ','.join(e.strip() for e in eans.split(','))  # Normalizar EANs
    
    # Verificar si el producto ya existe
    cursor.execute('SELECT eans FROM productos WHERE sku = ?', (sku,))
    result = cursor.fetchone()
    if result:
        # Actualizar EANs si ya existe
        existing_eans = result[0]
        new_eans = ','.join(set(existing_eans.split(',') + eans_combined.split(',')))
        cursor.execute('UPDATE productos SET eans = ? WHERE sku = ?', (new_eans, sku))
    else:
        # Insertar nuevo producto
        cursor.execute('INSERT INTO productos (sku, titulo, eans) VALUES (?, ?, ?)', (sku, titulo, eans_combined))
    
    print(f'Progreso: {idx}/{total_rows} productos migrados')

# Guardar cambios y cerrar la conexión
conn.commit()
conn.close()
print('Migración completada.')
