# Kivy Core Imports
import kivy
from kivy.app import App
from kivy.core.window import Window
from kivy.clock import Clock
from kivy.logger import Logger
from kivy.graphics import Color, Rectangle

# Kivy UI Components
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.checkbox import CheckBox
from kivy.uix.button import Button
from kivy.uix.slider import Slider
from kivy.uix.progressbar import ProgressBar
from kivy.uix.popup import Popup
from kivy.uix.switch import Switch
from kivy.uix.tabbedpanel import TabbedPanel, TabbedPanelItem
from kivy.uix.dropdown import DropDown
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.filechooser import FileChooserIconView, FileChooserListView
from kivy.uix.togglebutton import ToggleButton
from kivy.uix.scrollview import ScrollView
from kivy.animation import Animation

# Standard Library Imports
import os
import sqlite3
import logging
from datetime import datetime

# Third-Party Library Imports
from openpyxl import load_workbook, Workbook
import psutil

# Configurar el nivel de registro para ocultar mensajes de error específicos
Logger.setLevel(logging.CRITICAL)


class CustomFileChooser(FileChooserIconView):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.filters = [self.filter_hidden_files]  # Filtro personalizado para excluir archivos ocultos y protegidos

    def filter_hidden_files(self, folder, filename):
        try:
            # Ignorar archivos ocultos o protegidos del sistema
            protected_files = ['hiberfil.sys', 'pagefile.sys', 'swapfile.sys', 'dumpstack.log.tmp']
            return not filename.startswith('.') and filename.lower() not in protected_files
        except Exception:
            # Si ocurre un error al intentar acceder al archivo, ignorarlo
            return False

# Configuración de la ventana
Window.clearcolor = (0.1, 0.1, 0.1, 1)  # Fondo negro
Window.size = (650, 399)  # Tamaño inicial de la ventana
INITIAL_FONT_SIZE = 13    # Tamaño inicial de la fuente

# Variable para activar/desactivar el control de usuario/contraseña
ENABLE_LOGIN = False

# Variables para configurar el título dinámico
ENABLE_DYNAMIC_TITLE = True  # Activar o desactivar el título dinámico
TITLE_UPDATE_INTERVAL = 3  # Intervalo de actualización en segundos

# Función para actualizar el título de la ventana dinámicamente
def update_window_title(dt=None):
    """Actualiza el título de la ventana con el conteo dinámico y el uso de recursos."""
    if not ENABLE_DYNAMIC_TITLE:
        return  # Salir si el título dinámico está desactivado

    fecha = datetime.now().strftime('%d-%m-%Y')
    archivo = f'REVs/REV-{fecha}.xlsx'
    rev_count = 0
    ryt_count = 0

    if os.path.exists(archivo):
        wb = load_workbook(archivo, data_only=True)  # Asegurarse de cargar valores calculados
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row:  # Verificar que la fila no esté vacía
                estado = row[9]  # Columna "Estado"
                if estado == "Solo Revisión":
                    rev_count += 1
                elif estado == "Revisado y Traducido":
                    rev_count += 1
                    ryt_count += 1

    # Obtener el uso de CPU y RAM
    cpu_usage = psutil.cpu_percent(interval=0.1)
    ram_usage = psutil.virtual_memory().percent

    # Actualizar el título de la ventana
    Window.set_title(f'Contador de Revisiones V2.1 (OFICIAL) REV: {rev_count} / RYT: {ryt_count} (CPU: {cpu_usage}% / RAM: {ram_usage}%)')

class CustomSwitch(Switch):
    def __init__(self, **kwargs):
        super(CustomSwitch, self).__init__(**kwargs)
        self.bind(active=self.on_active)
        self.bind(pos=self.on_pos)
        self.bind(size=self.on_size)

    def on_active(self, instance, value):
        if self.canvas is not None:
            self.canvas.before.clear()
            with self.canvas.before:
                Color(0, 1, 0, 1) if value else Color(1, 0, 0, 1)
                Rectangle(pos=self.pos, size=self.size)

    def on_size(self, *args):
        self.on_active(self, self.active)

    def on_pos(self, *args):
        self.on_active(self, self.active)

class ContadorApp(App):
    # --- Configuración de paginación de resultados ---
    RESULTS_BLOCK_SIZE = 50  # Cambia este valor para modificar el tamaño del bloque de resultados por página

    def build(self):
        self.title = 'Contador de Revisiones V2.1 (OFICIAL)'
        self.screen_manager = ScreenManager()

        self.main_screen = Screen(name='main')
        self.main_screen.add_widget(self.build_main_interface())

        self.screen_manager.add_widget(self.main_screen)

        self.screen_manager.current = 'main'

        # Configurar el título dinámico de la ventana si está habilitado
        if ENABLE_DYNAMIC_TITLE:
            Clock.schedule_interval(update_window_title, TITLE_UPDATE_INTERVAL)

        Window.bind(on_resize=self.on_window_resize)

        return self.screen_manager

    def build_main_interface(self):
        self.root = BoxLayout(orientation='vertical', padding=10, spacing=10)
        Window.bind(on_resize=self.on_window_resize)
        self.init_db()
        Window.bind(on_request_close=self.on_request_close)
        
        # Botones superiores
        top_button_layout = BoxLayout(size_hint=(1, 0.1))
        self.historial_btn = Button(text='Historial', size_hint=(1, 1))
        self.historial_btn.bind(on_press=self.on_historial)
        self.reset_btn = Button(text='RESET!!!', size_hint=(1, 1))
        self.reset_btn.bind(on_press=self.on_reset)
        self.reset_btn.bind(on_release=self.on_reset_release)
        self.reg_db_btn = Button(text='+DB / MASS+', size_hint=(1, 1))  # Texto actualizado
        self.reg_db_btn.bind(on_release=self.on_reg_db_release)
        self.reg_db_btn.bind(on_press=self.on_reg_db_press)
        top_button_layout.add_widget(self.historial_btn)
        top_button_layout.add_widget(self.reset_btn)
        top_button_layout.add_widget(self.reg_db_btn)
        self.root.add_widget(top_button_layout)
        
        # Checkboxes para ZZ, LOTE, Set & Pack, Consumo, EDT & EDP, MakeUP
        self.check_zz = CheckBox(size_hint=(None, None))
        self.check_lote = CheckBox(size_hint=(None, None))
        self.check_set_pack = CheckBox(size_hint=(None, None))
        self.check_consumo = CheckBox(size_hint=(None, None))
        self.check_edt_edp = CheckBox(size_hint=(None, None))
        self.check_makeup = CheckBox(size_hint=(None, None))
        
        self.check_zz.bind(active=self.on_special_checkbox_active)
        self.check_lote.bind(active=self.on_special_checkbox_active)
        self.check_set_pack.bind(active=self.on_special_checkbox_active)
        self.check_consumo.bind(active=self.on_special_checkbox_active)
        self.check_edt_edp.bind(active=self.on_special_checkbox_active)
        self.check_makeup.bind(active=self.on_special_checkbox_active)
        
        special_checkbox_layout1 = BoxLayout(size_hint=(1, 0.1))
        special_checkbox_layout1.add_widget(Label(text='ZZ', color=(1, 1, 1, 1)))
        special_checkbox_layout1.add_widget(self.check_zz)
        special_checkbox_layout1.add_widget(Label(text='LOTE', color=(1, 1, 1, 1)))
        special_checkbox_layout1.add_widget(self.check_lote)
        special_checkbox_layout1.add_widget(Label(text='Set & Pack', color=(1, 1, 1, 1)))
        special_checkbox_layout1.add_widget(self.check_set_pack)
        
        special_checkbox_layout2 = BoxLayout(size_hint=(1, 0.1))
        special_checkbox_layout2.add_widget(Label(text='Consumo', color=(1, 1, 1, 1)))
        special_checkbox_layout2.add_widget(self.check_consumo)
        special_checkbox_layout2.add_widget(Label(text='MakeUP', color=(1, 1, 1, 1)))
        special_checkbox_layout2.add_widget(self.check_makeup)
        special_checkbox_layout2.add_widget(Label(text='EDT & EDP', color=(1, 1, 1, 1)))
        special_checkbox_layout2.add_widget(self.check_edt_edp)
        
        self.root.add_widget(special_checkbox_layout1)
        self.root.add_widget(special_checkbox_layout2)
        
        # Combobox para seleccionar tipo
        self.dropdown = DropDown()
        self.tipo_combobox = Button(text='Seleccionar Tipo', size_hint=(0.5, 1))
        self.tipo_combobox.bind(on_release=self.open_dropdown)
        
        # Lista de tipos
        self.tipos = ['ACCESSORIES', 'administración', 'AFEITAR AFTER PRESHAVE', 'AFEITAR CREMA ESPUMA BROCHA', 'AFEITAR HOJA MAQUINA BROCHA', 'ALIMENTOS ENVASADOS', 'ALIMENTOS MASCOTAS', 'ALMACEN VARIOS', 'AMBIENTADORES', 'ANEXOS', 'Automatico desde Articulo', 'BAÑO DESODORANTE', 'BAÑO GEL', 'BAÑO JABON', 'BAÑO LECHE BODY L. ACEITE', 'BAÑO TALCO', 'BAÑO-VARIOS', 'BEBIDAS ENVASADAS', 'BEELINE', 'BISUTERIA', 'CABELLO ACONDIC. SUAVIZANTE', 'CABELLO CHAMPU', 'CABELLO FIJADOR BRILLANTINA', 'CABELLO LACA', 'CABELLO TINTES', 'CABELLO TONICO LOCION', 'CHRISTMAS', 'DENTIFRICO', 'DEPILATORIO', 'DESCUENTO PROMO', 'DUPLOS', 'ESTUCHES COLORIDO FLORES', 'ESTUCHES TRATAMIENTO', 'GIFT WRAPPING', 'GRANELES', 'HIGIENE CELULOSA', 'HIGIENE MASCOTAS', 'HOBBY', 'HOME INTERIOR', 'JUEGOS EROTICOS', 'LOTES', 'MAQUILLAJE CUERPO', 'MAQUILLAJE LABIOS', 'MAQUILLAJE MANOS', 'MAQUILLAJE OJOS', 'MAQUILLAJE ROSTRO', 'MAQUILLAJE SURTIDO', 'MATERIAL CONSUMIBLE', 'MATERIAL PLV', 'MATERIAL PLV ESPECIFICO', 'MINIATURAS', 'MUY MUCHO', 'P01', 'PARTY ARTICLES', 'PELUCHES JUGUETES', 'PELUQUERIA FRANCK PROVOST', 'PERF. ESTUCHES HOMBRE', 'PERF. ESTUCHES MUJER', 'PERF.ALC.FEMENINA', 'PERF.ALC.FEMENINA ALMACEN', 'PERF.ALC.INFANTIL', 'PERF.ALC.INFANTIL ALMACEN', 'PERF.ALC.MASCULINA', 'PERF.ALC.MASCULINA ALMACEN', 'PROMOCIONAL FEMENINO ALMACEN', 'PROMOCIONAL MASCULINO ALMACEN', 'PROMOCIONALES FEMENINOS', 'PROMOCIONALES MASCULINOS', 'PRUEBAS EXCEL', 'SEASON', 'STATIONERY', 'SUSCRIPCIONES', 'TARJETAS REGALO DIGITALES', 'TARJETAS REGALO FISICAS', 'TEENS', 'TEXTIL', 'TOYS', 'TRAT.FEMENINO', 'TRAT.MASCULINO', 'TRAT.SOLAR', 'TRATAMIENTO CUERPO MANOS', 'VALE', 'VARIOS', 'VARIOS SIN CODIFICAR', 'VARIOUS ITEMS']
        for tipo in self.tipos:
            btn = Button(text=tipo, size_hint_y=None, height=44)
            btn.bind(on_release=lambda btn: self.dropdown.select(btn.text))
            self.dropdown.add_widget(btn)
        self.dropdown.bind(on_select=self.on_tipo_select)
        
        # Botones EX1 y EX2
        self.ex1_btn = Button(text='LOCK', size_hint=(0.25, 1))
        self.ex1_btn.bind(on_press=self.toggle_lock_mode)
        self.ex2_btn = Button(text='F4K3', size_hint=(0.25, 1))
        self.ex2_btn.bind(on_press=self.on_f4k3_press)  # Asignar función al botón F4K3

        combobox_layout = BoxLayout(size_hint=(1, 0.1))
        combobox_layout.add_widget(self.ex1_btn)
        combobox_layout.add_widget(self.tipo_combobox)
        combobox_layout.add_widget(self.ex2_btn)
        
        self.root.add_widget(combobox_layout)
        
        # Campo de texto EAN/SKU/ID
        self.ean_sku_id = TextInput(hint_text='EAN/SKU/ID', multiline=False, size_hint=(1, 0.1))
        self.ean_sku_id.bind(on_text_validate=self.on_ean_enter)
        self.root.add_widget(self.ean_sku_id)
        
        # Campo de texto MARCA/TITULO
        self.marca_titulo = TextInput(hint_text='MARCA/TITULO', multiline=False, size_hint=(1, 0.1))
        self.marca_titulo.bind(on_text_validate=self.on_marca_titulo_enter)
        self.marca_titulo.bind(on_text_validate=self.focus_next)
        self.root.add_widget(self.marca_titulo)
        
        # Checkboxes
        self.check_pt = CheckBox(size_hint=(None, None), size=(48, 48))
        self.check_es = CheckBox(size_hint=(None, None), size=(48, 48))
        self.check_it = CheckBox(size_hint=(None, None), size=(48, 48))
        
        checkbox_layout = BoxLayout(size_hint=(1, 0.1))
        checkbox_layout.add_widget(Label(text='Tiene PT', color=(1, 1, 1, 1)))
        checkbox_layout.add_widget(self.check_pt)
        checkbox_layout.add_widget(Label(text='Tiene ES', color=(1, 1, 1, 1)))
        checkbox_layout.add_widget(self.check_es)
        checkbox_layout.add_widget(Label(text='Tiene IT', color=(1, 1, 1, 1)))
        checkbox_layout.add_widget(self.check_it)
        self.root.add_widget(checkbox_layout)
        
        # Campo numérico
        self.slider_value = TextInput(text='1', multiline=False, size_hint=(None, None), halign='center')  # Valor inicial cambiado a 1

        # El layout solo tendrá el campo numérico centrado
        slider_layout = BoxLayout(size_hint=(1, 0.1))
        slider_layout.add_widget(Label(size_hint=(0.4, 1)))  # Espacio izquierdo
        slider_layout.add_widget(self.slider_value)
        slider_layout.add_widget(Label(size_hint=(0.4, 1)))  # Espacio derecho
        self.root.add_widget(slider_layout)
        
        self.slider_value.bind(on_text_validate=self.on_text_value_change)
        
        # Checkboxes para UND, ML, GR
        self.check_und = CheckBox(size_hint=(None, None), size=(48, 48))
        self.check_ml = CheckBox(size_hint=(None, None), size=(48, 48))
        self.check_gr = CheckBox(size_hint=(None, None), size=(48, 48))
        
        self.check_und.bind(active=self.on_unit_checkbox_active)
        self.check_ml.bind(active=self.on_unit_checkbox_active)
        self.check_gr.bind(active=self.on_unit_checkbox_active)
        
        unit_checkbox_layout = BoxLayout(size_hint=(1, 0.1))
        unit_checkbox_layout.add_widget(Label(text='UND', color=(1, 1, 1, 1)))
        unit_checkbox_layout.add_widget(self.check_und)
        unit_checkbox_layout.add_widget(Label(text='ML', color=(1, 1, 1, 1)))
        unit_checkbox_layout.add_widget(self.check_ml)
        unit_checkbox_layout.add_widget(Label(text='GR', color=(1, 1, 1, 1)))
        unit_checkbox_layout.add_widget(self.check_gr)
        self.root.add_widget(unit_checkbox_layout)
        
        # Botones
        self.revisado_btn = Button(text='REVISADO', size_hint=(1, 1))
        self.revisado_btn.bind(on_press=self.on_revisado)
        self.traducir_btn = Button(text='TRADUCIR', size_hint=(1, 1))
        self.traducir_btn.bind(on_press=self.on_traducir)
        self.traducido_btn = Button(text='TRADUCIDO', size_hint=(1, 1))
        self.traducido_btn.bind(on_press=self.on_traducido)
        
        button_layout = BoxLayout(size_hint=(1, 0.1))
        button_layout.add_widget(self.revisado_btn)
        button_layout.add_widget(self.traducir_btn)
        button_layout.add_widget(self.traducido_btn)
        self.root.add_widget(button_layout)
        
        # Barra de estado
        self.status_bar = Label(text='Estado: Esperando...', size_hint=(1, 0.1), color=(1, 1, 1, 1))
        self.status_bar.bind(on_touch_down=self.on_status_bar_double_click)
        self.root.add_widget(self.status_bar)
        
        self.descripcion = ''
        self.modo_empleo = ''
        self.precauciones = ''
        self.mas_informaciones = ''
        self.traduccion_tipo = ''
        
        self.descripcion_pt = ''
        self.modo_empleo_pt = ''
        self.precauciones_pt = ''
        self.mas_informaciones_pt = ''
        self.descripcion_it = ''
        self.modo_empleo_it = ''
        self.precauciones_it = ''
        self.mas_informaciones_it = ''
        
        self.lock_mode = False  # Estado inicial del modo bloqueo
        self.locked_values = {}  # Diccionario para almacenar los valores bloqueados
        
        Clock.schedule_interval(self.check_focus, 0.1)  # Verificar el foco cada 0.1 segundos
        Clock.schedule_once(lambda dt: self.update_all_fonts(), 0)
        return self.root

    def init_db(self):
        self.conn = sqlite3.connect('db.db')
        self.cursor = self.conn.cursor()
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS productos (
                sku TEXT PRIMARY KEY,
                titulo TEXT,
                eans TEXT
            )
        ''')
        self.conn.commit()

    def on_window_resize(self, instance, width, height):
        if (width, height) == (580, 391):
            base_font = INITIAL_FONT_SIZE
        else:
            base_font = int(min(Window.width, Window.height) * 0.035)
        self.status_bar.text = f'Estado: Ventana redimensionada a {width}x{height} | Fuente: {base_font}'
        self.update_all_fonts()

    def update_all_fonts(self):
        # Calcula un tamaño base de fuente según el tamaño de la ventana o usa el inicial si es el tamaño inicial
        if Window.size == (580, 391):
            base_font = INITIAL_FONT_SIZE
        else:
            base_font = int(min(Window.width, Window.height) * 0.035)
        small_font = int(base_font * 0.8)
        big_font = int(base_font * 1.2)

        def update_widget_font(widget):
            # Ajusta el tamaño de fuente de los widgets relevantes
            if isinstance(widget, (Label, Button, TextInput)):
                widget.font_size = base_font
                # Ajusta el tamaño del TextInput numérico
                if widget is self.slider_value:
                    widget.width = Window.width * 0.12
                    widget.height = Window.height * 0.08
            if isinstance(widget, CheckBox):
                widget.size = (Window.height * 0.08, Window.height * 0.08)
            # Recorre hijos si es un layout
            if hasattr(widget, 'children'):
                for child in widget.children:
                    update_widget_font(child)

        update_widget_font(self.root)
        # Ajusta popups si están abiertos
        for attr in ['loading_popup', 'info_popup', 'add_product_popup', 'add_to_db_popup', 'progress_popup', 'results_popup', 'missing_products_popup', 'import_confirmation_popup', 'progress_overlay', 'traducir_popup', 'historial_popup', 'exit_confirmation_popup']:
            popup = getattr(self, attr, None)
            if popup and hasattr(popup, 'content'):
                update_widget_font(popup.content)

    def focus_next(self, instance):
        next_widget = instance.get_focus_next()
        if (next_widget):
            next_widget.focus = True

    def on_slider_value_change(self, instance, value):
        # Eliminado porque ya no hay slider
        pass

    def on_text_value_change(self, instance):
        value = self.slider_value.text
        if value.isdigit() and 0 <= int(value) <= 1000:
            # self.slider.value = int(value)  # Ya no hay slider
            self.status_bar.text = f'Estado: Valor del campo numérico cambiado a {value}'
        else:
            self.slider_value.text = '1'

    def on_ean_enter(self, instance):
        ean = self.ean_sku_id.text.strip()
        if not ean:
            self.show_warning_popup('El campo EAN/SKU/ID\nno puede estar vacío.')
            return

        self.show_loading_popup('Espere, cargando...')
        self.root.do_layout()
        results = self.search_product_in_db(ean)
        self.loading_popup.dismiss()
        if results:
            if len(results) == 1:
                sku, title = results[0]
                revision_status = self.check_revision_status(sku)
                self.ean_sku_id.text = sku
                self.marca_titulo.text = title
                self.show_info_popup('Producto encontrado', f'SKU: {sku}\nTítulo: {title}\n{revision_status}')
            else:
                # Si hay más de un resultado, mostrar popup para elegir
                self.show_results_popup(results)
        else:
            self.show_add_product_popup(ean)
        self.ean_sku_id.focus = True  # Asegurar el foco en el campo "EAN/SKU/ID"

    def search_product_in_db(self, ean):
        """
        Busca productos por SKU exacto o por EAN exacto (en la lista de EANs separados por coma).
        Permite valores como NO-EAN y NO-DESC.
        Devuelve una lista de tuplas (sku, titulo) de los productos encontrados.
        """
        try:
            # Buscar por SKU exacto (clave única)
            self.cursor.execute('SELECT sku, titulo FROM productos WHERE sku = ?', (ean,))
            row = self.cursor.fetchone()
            if row:
                return [row]

            # Buscar por EAN exacto en la lista de EANs (separados por coma)
            self.cursor.execute('SELECT sku, titulo, eans FROM productos')
            matches = []
            for sku, titulo, eans in self.cursor.fetchall():
                if eans:
                    ean_list = [x.strip() for x in eans.split(',')]
                    if ean in ean_list:
                        matches.append((sku, titulo))
            return matches
        except sqlite3.OperationalError as e:
            if "database is locked" in str(e):
                msg = "La base de datos está en uso por otro proceso. Por favor, cierre cualquier programa que esté usando 'db.db' y vuelva a intentarlo."
                print(msg)
                self.show_warning_popup(msg)
                return []
            else:
                raise

    def check_revision_status(self, sku):
        fecha = datetime.now().strftime('%d-%m-%Y')
        archivo = f'REVs/REV-{fecha}.xlsx'
        if os.path.exists(archivo):
            try:
                wb = load_workbook(archivo)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] == sku:
                        return 'YA REVISADO/TRADUCIDO'
            except Exception as e:
                print(f"Error al leer el archivo de revisiones: {e}")
        return 'SIN REVISION'

    def show_loading_popup(self, message):
        content = BoxLayout(orientation='vertical', padding=10)
        content.add_widget(Label(text=message, text_size=(280, None), halign='center'))
        self.loading_popup = Popup(title='Cargando',
                                   content=content,
                                   size_hint=(0.6, 0.4),
                                   auto_dismiss=False)
        self.loading_popup.open()

    def show_info_popup(self, title, message):
        content = BoxLayout(orientation='vertical', padding=10)
        content.add_widget(Label(text=message, text_size=(280, None), halign='center'))
        popup = Popup(title=title,
                      content=content,
                      size_hint=(0.6, 0.4),
                      auto_dismiss=False)
        popup.bind(on_dismiss=self.on_info_popup_dismiss)
        Window.bind(on_key_down=self.on_key_down)
        self.info_popup = popup
        popup.open()

    def on_info_popup_dismiss(self, instance):
        Window.unbind(on_key_down=self.on_key_down)  # Corregido: especificar el evento y el método a desregistrar

    def on_key_down(self, window, key, *args):
        if key in [27, 13]:  # Códigos de tecla ESC [27] y ENTER [13]
            if hasattr(self, 'info_popup') and self.info_popup:
                self.info_popup.dismiss()
                return True
        return False

    def show_add_product_popup(self, ean):
        content = BoxLayout(orientation='vertical', padding=10)
        content.add_widget(Label(text=f'El producto con EAN {ean} no se encontró.'))
        content.add_widget(Label(text='¿Desea agregarlo a la base de datos?'))
        button_layout = BoxLayout(size_hint=(1, 0.2))
        continue_button = Button(text='Continuar')
        continue_button.bind(on_press=lambda x: self.continue_without_adding(ean))
        add_button = Button(text='Añadir a DB')
        add_button.bind(on_press=lambda x: self.show_add_to_db_popup(ean))
        button_layout.add_widget(continue_button)
        button_layout.add_widget(add_button)
        content.add_widget(button_layout)
        self.add_product_popup = Popup(title='Agregar Producto',
                                       content=content,
                                       size_hint=(0.8, 0.4))
        self.add_product_popup.open()

    def continue_without_adding(self, ean):
        self.add_product_popup.dismiss()
        self.ean_sku_id.text = ean

    def show_add_to_db_popup(self, instance=None):
        content = BoxLayout(orientation='vertical', padding=10, spacing=10)
        self.sku_input = TextInput(hint_text='SKU (Codigo ID Unico)', multiline=False)
        self.title_input = TextInput(hint_text='Titulo (Marca y Descripción)', multiline=False)
        self.eans_input = TextInput(hint_text='EANs (separados por coma si hay varios)', multiline=False)
        add_button = Button(text='Añadir')
        add_button.bind(on_press=self.add_product_to_db)
        content.add_widget(self.sku_input)
        content.add_widget(self.title_input)
        content.add_widget(self.eans_input)
        content.add_widget(add_button)
        self.add_to_db_popup = Popup(title='Añadir a la Base de Datos',
                                     content=content,
                                     size_hint=(0.8, 0.6))
        self.add_to_db_popup.open()

    def add_product_to_db(self, instance):
        sku = self.sku_input.text.strip()
        title = self.title_input.text.strip()
        eans = self.eans_input.text.strip()
        if sku and title and eans:
            self.show_loading_popup('Añadiendo a la base de datos...')
            self.root.do_layout()
            try:
                self.cursor.execute('INSERT INTO productos (sku, titulo, eans) VALUES (?, ?, ?)', (sku, title, eans))
                self.conn.commit()
                self.loading_popup.dismiss()
                if hasattr(self, 'add_to_db_popup'):  # Verificar si el popup existe
                    self.add_to_db_popup.dismiss()
                self.ean_sku_id.text = sku
                self.marca_titulo.text = title
            except sqlite3.OperationalError as e:
                self.loading_popup.dismiss()
                if "database is locked" in str(e):
                    msg = "La base de datos está en uso por otro proceso. Por favor, cierre cualquier programa que esté usando 'db.db' y vuelva a intentarlo."
                    print(msg)
                    self.show_warning_popup(msg)
                else:
                    raise
        else:
            self.show_warning_popup('Todos los campos son obligatorios.')

    def on_tipo_select(self, instance, x):
        self.tipo_combobox.text = x
        self.selected_tipo = x

    def on_special_checkbox_active(self, checkbox, value):
        if value:
            if checkbox == self.check_zz:
                self.check_lote.active = False
                self.check_set_pack.active = False
                self.check_consumo.active = False
                self.check_edt_edp.active = False
                self.check_makeup.active = False
            elif checkbox == self.check_lote:
                self.check_zz.active = False
                self.check_set_pack.active = False
                self.check_consumo.active = False
                self.check_edt_edp.active = False
                self.check_makeup.active = False
                self.show_lote_popup()
            elif checkbox == self.check_set_pack:
                self.check_zz.active = False
                self.check_lote.active = False
                self.check_consumo.active = False
                self.check_edt_edp.active = False
                self.check_makeup.active = False
                self.show_lote_popup()
            elif checkbox == self.check_consumo:
                self.check_zz.active = False
                self.check_lote.active = False
                self.check_set_pack.active = False
                self.check_edt_edp.active = False
                self.check_makeup.active = False
            elif checkbox == self.check_edt_edp:
                self.check_zz.active = False
                self.check_lote.active = False
                self.check_set_pack.active = False
                self.check_consumo.active = False
                self.check_makeup.active = False
            elif checkbox == self.check_makeup:
                self.check_zz.active = False
                self.check_lote.active = False
                self.check_set_pack.active = False
                self.check_consumo.active = False
                self.check_edt_edp.active = False

    def show_lote_popup(self):
        self.lote_popup = Popup(title='Composición de Lote',
                                content=BoxLayout(orientation='vertical'),
                                size_hint=(0.8, 0.5),
                                auto_dismiss=False)  # Evitar que el popup se cierre al hacer clic fuera
        self.lote_text_input = TextInput(hint_text='EANs separados por líneas', multiline=True)
        next_button = Button(text='Siguiente', size_hint=(1, 0.2))
        next_button.bind(on_press=self.on_next_lote)
        accept_button = Button(text='Aceptar', size_hint=(1, 0.2))
        accept_button.bind(on_press=self.on_accept_lote)
        
        self.lote_popup.content.add_widget(self.lote_text_input)
        self.lote_popup.content.add_widget(next_button)
        self.lote_popup.content.add_widget(accept_button)
        self.lote_popup.open()

    def on_next_lote(self, instance):
        self.lote_text_input.text += '\n'

    def on_accept_lote(self, instance):
        eans = self.lote_text_input.text.strip().split('\n')
        self.lote_composition = ','.join([f'"{ean.strip()}"' for ean in eans if ean.strip()])
        self.lote_popup.dismiss()

    def on_unit_checkbox_active(self, checkbox, value):
        if value:
            if checkbox == self.check_und:
                self.check_ml.active = False
                self.check_gr.active = False
            elif checkbox == self.check_ml:
                self.check_und.active = False
                self.check_gr.active = False
            elif checkbox == self.check_gr:
                self.check_und.active = False
                self.check_ml.active = False

    def show_warning_popup(self, message):
        content = BoxLayout(orientation='vertical', padding=10)
        label = Label(text=message, text_size=(280, None), halign='center', valign='middle', size_hint_y=None)
        label.bind(texture_size=label.setter('size'))
        content.add_widget(label)
        popup = Popup(title='Advertencia',
                      content=content,
                      size_hint=(0.6, 0.4))
        popup.open()

    def on_revisado(self, instance):
        if not self.ean_sku_id.text.strip():
            self.show_warning_popup('El campo EAN/SKU/ID\nno puede estar vacío.')
        else:
            if self.lock_mode:
                self.apply_locked_values()
            self.registrar_revision('Solo Revisión')
            self.status_bar.text = 'Estado: Producto revisado'
            self.reset_fields()  # Limpiar campos después de revisar
            self.ean_sku_id.focus = True  # Asegurar el foco en el campo "EAN/SKU/ID"

    def on_traducir(self, instance):
        self.traducir_popup = Popup(title='Traducciones',
                                    content=BoxLayout(orientation='vertical', spacing=10, padding=10),
                                    size_hint=(0.8, 0.8),
                                    auto_dismiss=False)  # Evitar que el popup se cierre al hacer clic fuera

        tab_panel = TabbedPanel(do_default_tab=False)
        
        pt_tab = TabbedPanelItem(text='PT')
        it_tab = TabbedPanelItem(text='IT')

        self.descripcion_input_pt = TextInput(hint_text='Descripcion', multiline=True, size_hint=(1, 0.2))
        self.modo_empleo_input_pt = TextInput(hint_text='Modo de Empleo', multiline=True, size_hint=(1, 0.2))
        self.precauciones_input_pt = TextInput(hint_text='Precauciones', multiline=True, size_hint=(1, 0.2))
        self.mas_informaciones_input_pt = TextInput(hint_text='Más Informaciones', multiline=True, size_hint=(1, 0.2))

        self.descripcion_input_it = TextInput(hint_text='Descripcion', multiline=True, size_hint=(1, 0.2))
        self.modo_empleo_input_it = TextInput(hint_text='Modo de Empleo', multiline=True, size_hint=(1, 0.2))
        self.precauciones_input_it = TextInput(hint_text='Precauciones', multiline=True, size_hint=(1, 0.2))
        self.mas_informaciones_input_it = TextInput(hint_text='Más Informaciones', multiline=True, size_hint=(1, 0.2))

        pt_tab_content = BoxLayout(orientation='vertical')
        pt_tab_content.add_widget(self.descripcion_input_pt)
        pt_tab_content.add_widget(self.modo_empleo_input_pt)
        pt_tab_content.add_widget(self.precauciones_input_pt)
        pt_tab_content.add_widget(self.mas_informaciones_input_pt)
        pt_tab.add_widget(pt_tab_content)

        it_tab_content = BoxLayout(orientation='vertical')
        it_tab_content.add_widget(self.descripcion_input_it)
        it_tab_content.add_widget(self.modo_empleo_input_it)
        it_tab_content.add_widget(self.precauciones_input_it)
        it_tab_content.add_widget(self.mas_informaciones_input_it)
        it_tab.add_widget(it_tab_content)

        tab_panel.add_widget(pt_tab)
        tab_panel.add_widget(it_tab)

        grabar_button = Button(text='Grabar y Volver', size_hint=(1, 0.2))
        grabar_button.bind(on_press=self.on_grabar_traducciones)

        self.traducir_popup.content.add_widget(tab_panel)
        self.traducir_popup.content.add_widget(grabar_button)
        self.traducir_popup.open()

        self.load_traduccion_data()

    def save_traduccion_data(self):
        self.descripcion_pt = self.descripcion_input_pt.text
        self.modo_empleo_pt = self.modo_empleo_input_pt.text
        self.precauciones_pt = self.precauciones_input_pt.text
        self.mas_informaciones_pt = self.mas_informaciones_input_pt.text

        self.descripcion_it = self.descripcion_input_it.text
        self.modo_empleo_it = self.modo_empleo_input_it.text
        self.precauciones_it = self.precauciones_input_it.text
        self.mas_informaciones_it = self.mas_informaciones_input_it.text

    def load_traduccion_data(self):
        self.descripcion_input_pt.text = self.descripcion_pt
        self.modo_empleo_input_pt.text = self.modo_empleo_pt
        self.precauciones_input_pt.text = self.precauciones_pt
        self.mas_informaciones_input_pt.text = self.mas_informaciones_pt

        self.descripcion_input_it.text = self.descripcion_it
        self.modo_empleo_input_it.text = self.modo_empleo_it
        self.precauciones_input_it.text = self.precauciones_it
        self.mas_informaciones_input_it.text = self.mas_informaciones_it

    def on_grabar_traducciones(self, instance):
        self.save_traduccion_data()
        self.traducir_popup.dismiss()

    def on_traducido(self, instance):
        if not self.ean_sku_id.text.strip():
            self.show_warning_popup('El campo EAN/SKU/ID\nno puede estar vacío.')
        else:
            if self.lock_mode:
                self.apply_locked_values()
            self.registrar_revision('Revisado y Traducido')
            self.status_bar.text = 'Estado: Producto traducido'
            self.reset_fields()  # Limpiar campos después de traducir
            self.ean_sku_id.focus = True  # Asegurar el foco en el campo "EAN/SKU/ID"

    def registrar_revision(self, estado):
        ean_sku_id = self.ean_sku_id.text
        marca_titulo = self.marca_titulo.text
        tipo = self.selected_tipo if hasattr(self, 'selected_tipo') else 'ZZ' if self.check_zz.active else 'LOTE' if self.check_lote.active else 'Set & Pack' if self.check_set_pack.active else 'Consumo' if self.check_consumo.active else 'EDT & EDP' if self.check_edt_edp.active else 'MakeUP' if self.check_makeup.active else ''
        tiene_pt = 'Tiene PT' if self.check_pt.active else 'No Tiene PT - TRADUZIDO'
        tiene_es = 'Tiene ES' if self.check_es.active else 'No Tiene ES - TRADUCIDO'
        tiene_it = 'Tiene IT' if self.check_it.active else 'No Tiene IT - TRADOTTO'
        cantidad_neta = self.slider_value.text
        unidad = 'UND' if self.check_und.active else 'ML' if self.check_ml.active else 'GR' if self.check_gr.active else ''
        composicion_lote = self.lote_composition if self.check_lote.active or self.check_set_pack.active else ''
        
        fecha = datetime.now().strftime('%d-%m-%Y')
        archivo = f'REVs/REV-{fecha}.xlsx'
        
        descripcion_col = f'Descripcion{self.traduccion_tipo}' if self.traduccion_tipo else ''
        modo_empleo_col = f'Modo de Empleo{self.traduccion_tipo}' if self.traduccion_tipo else ''
        precauciones_col = f'Precauciones{self.traduccion_tipo}' if self.traduccion_tipo else ''
        mas_informaciones_col = f'Más Informaciones{self.traduccion_tipo}' if self.traduccion_tipo else ''
        
        if not os.path.exists('REVs'):
            os.makedirs('REVs')
        
        try:
            if os.path.exists(archivo):
                wb = load_workbook(archivo)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(['EAN/SKU/ID', 'MARCA/TITULO', 'Tipo', 'Tiene PT', 'Tiene ES', 'Tiene IT', 'Cantidad Neta', 'UND/ML/GR', 'Composición de Lote', 'Estado', 'DescripcionPT', 'Modo de EmpleoPT', 'PrecaucionesPT', 'Más InformacionesPT', 'DescripcionIT', 'Modo de EmpleoIT', 'PrecaucionesIT', 'Más InformacionesIT'])
            
            ws.append([ean_sku_id, marca_titulo, tipo, tiene_pt, tiene_es, tiene_it, cantidad_neta, unidad, composicion_lote, estado, self.descripcion_pt, self.modo_empleo_pt, self.precauciones_pt, self.mas_informaciones_pt, self.descripcion_it, self.modo_empleo_it, self.precauciones_it, self.mas_informaciones_it])
            try:
                wb.save(archivo)
            except PermissionError as e:
                msg = f"No se puede guardar el archivo de revisiones '{archivo}'. Es posible que esté abierto en otro programa. Por favor, ciérrelo y vuelva a intentarlo."
                print(msg)
                self.show_warning_popup(msg)
                return
            except Exception as e:
                msg = f"Error inesperado al guardar el archivo de revisiones: {e}"
                print(msg)
                self.show_warning_popup(msg)
                return

            self.reset_fields()  # Limpiar campos después de registrar la revisión
        except Exception as e:
            msg = f"Error al registrar la revisión: {e}"
            print(msg)
            self.show_warning_popup(msg)

    def reset_fields(self):
        self.ean_sku_id.text = ''
        self.marca_titulo.text = ''
        self.check_zz.active = False
        self.check_lote.active = False
        self.check_set_pack.active = False
        self.check_consumo.active = False
        self.check_edt_edp.active = False
        self.check_makeup.active = False
        self.check_pt.active = False
        self.check_es.active = False
        self.check_it.active = False
        self.slider_value.text = '1'  # Volver a 1 después de resetear la interfaz
        self.check_und.active = False
        self.check_ml.active = False
        self.check_gr.active = False
        self.lote_composition = ''
        self.descripcion_pt = ''
        self.modo_empleo_pt = ''
        self.precauciones_pt = ''
        self.mas_informaciones_pt = ''
        self.descripcion_it = ''
        self.modo_empleo_it = ''
        self.precauciones_it = ''
        self.mas_informaciones_it = ''
        if self.lock_mode:
            self.apply_locked_values()
        else:
            self.tipo_combobox.text = 'Seleccionar Tipo'
            if hasattr(self, 'selected_tipo'):
                del self.selected_tipo
        self.ean_sku_id.focus = True  # Asegurar el foco en el campo "EAN/SKU/ID"
        
        # Inicializar campos de entrada de traducción si no existen
        if hasattr(self, 'descripcion_input_pt'):
            self.load_traduccion_data()  # Limpiar los campos de traducción

    def focus_ean_sku_id(self, dt):
        self.ean_sku_id.focus = True

    def on_status_bar_double_click(self, instance, touch):
        if touch.is_double_tap:
            Window.size = (580, 391)
            # Esperar a que el evento de resize termine antes de actualizar fuentes y estado
            def restore_font_and_status(dt):
                base_font = INITIAL_FONT_SIZE
                self.update_all_fonts()
                self.status_bar.text = f'Estado: Ventana restablecida a tamaño inicial {Window.width}x{Window.height} | Fuente: {base_font}'
            Clock.schedule_once(restore_font_and_status, 0.05)

    def on_historial(self, instance):
        self.historial_popup = Popup(title='Historial de Revisiones',
                                     content=BoxLayout(orientation='vertical', padding=10, spacing=10),
                                     size_hint=(0.8, 0.8))
        self.historial_content = BoxLayout(orientation='vertical')
        self.historial_popup.content.add_widget(self.historial_content)
        
        button_layout = BoxLayout(size_hint=(1, 0.2))
        self.historial_volver_btn = Button(text='Volver')
        self.historial_volver_btn.bind(on_press=self.on_historial_volver)
        self.historial_siguiente_btn = Button(text='Siguiente')
        self.historial_siguiente_btn.bind(on_press=self.on_historial_siguiente)
        button_layout.add_widget(self.historial_volver_btn)
        button_layout.add_widget(self.historial_siguiente_btn)
        
        self.historial_popup.content.add_widget(button_layout)
        self.historial_index = 0
        self.load_historial()
        self.historial_popup.open()

    def load_historial(self):
        self.historial_content.clear_widgets()
        fecha = datetime.now().strftime('%d-%m-%Y')
        archivo = f'REVs/REV-{fecha}.xlsx'
        if os.path.exists(archivo):
            wb = load_workbook(archivo)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            rows.reverse()
            for row in rows[self.historial_index:self.historial_index + 5]:
                self.historial_content.add_widget(Label(text=f'{row[0]}-{row[1]}-{row[2]} / {row[9]}'))

    def on_historial_siguiente(self, instance):
        self.historial_index += 5
        self.load_historial()

    def on_historial_volver(self, instance):
        if self.historial_index >= 5:
            self.historial_index -= 5
        self.load_historial()

    def on_reset(self, instance):
        self.reset_start_time = datetime.now()
        self.status_bar.text = 'Estado: Mantenga presionado para resetear...'
        self.show_countdown_animation(instance)  # Mostrar la animación de cuenta regresiva
        Clock.schedule_once(self.reset_ready, 3)

    def show_countdown_animation(self, instance):
        """Muestra una animación de cuenta regresiva sobre la UI."""
        self.countdown_label = Label(text="3", font_size=50, color=(1, 0, 0, 1), size_hint=(None, None), size=(100, 100))
        # Agregar el widget al contenido de la pantalla actual
        current_screen = self.screen_manager.current_screen
        current_screen.add_widget(self.countdown_label)

        def update_countdown_label(dt):
            remaining_time = 3 - (datetime.now() - self.reset_start_time).total_seconds()
            if remaining_time > 0:
                self.countdown_label.text = str(int(remaining_time))
            else:
                current_screen.remove_widget(self.countdown_label)

        self.countdown_animation = Animation(x=instance.center_x, y=instance.center_y, duration=3)
        self.countdown_animation.bind(on_complete=lambda *args: current_screen.remove_widget(self.countdown_label))
        self.countdown_animation.start(self.countdown_label)
        Clock.schedule_interval(update_countdown_label, 0.1)

    def on_reset_release(self, instance):
        current_screen = self.screen_manager.current_screen
        if hasattr(self, 'countdown_label') and self.countdown_label in current_screen.children:
            current_screen.remove_widget(self.countdown_label)  # Eliminar la animación si se cancela
        if (datetime.now() - self.reset_start_time).total_seconds() >= 3:
            self.reset_fields()
            self.status_bar.text = 'Estado: Interfaz reseteada'
        else:
            self.status_bar.text = 'Estado: Reset cancelado'

    def reset_ready(self, dt):
        """Completa el proceso de reinicio después de la cuenta regresiva."""
        self.reset_fields()
        self.status_bar.text = 'Estado: Interfaz reseteada'

    def toggle_lock_mode(self, instance):
        self.lock_mode = not self.lock_mode
        if self.lock_mode:
            self.locked_values = {
                'tipo': self.tipo_combobox.text,
                'check_zz': self.check_zz.active,
                'check_lote': self.check_lote.active,
                'check_set_pack': self.check_set_pack.active,
                'check_consumo': self.check_consumo.active,
                'check_edt_edp': self.check_edt_edp.active,
                'check_makeup': self.check_makeup.active,
                'check_pt': self.check_pt.active,
                'check_es': self.check_es.active,
                'check_it': self.check_it.active,
                'slider_text': self.slider_value.text,  # Agregar el valor del campo numérico
                'check_und': self.check_und.active,
                'check_ml': self.check_ml.active,
                'check_gr': self.check_gr.active
            }
            self.ex1_btn.background_color = (1, 0, 0, 1)  # Cambiar color del botón a rojo
            self.status_bar.text = 'Estado: Modo bloqueo activado'
            self.ean_sku_id.focus = True  # Asegurar el foco en el campo "EAN/SKU/ID"
        else:
            self.ex1_btn.background_color = (1, 1, 1, 1)  # Restaurar color del botón
            self.status_bar.text = 'Estado: Modo bloqueo desactivado'

    def apply_locked_values(self):
        self.tipo_combobox.text = self.locked_values.get('tipo', '')
        self.check_zz.active = self.locked_values.get('check_zz', False)
        self.check_lote.active = self.locked_values.get('check_lote', False)
        self.check_set_pack.active = self.locked_values.get('check_set_pack', False)  # Corregir KeyError
        self.check_consumo.active = self.locked_values.get('check_consumo', False)
        self.check_edt_edp.active = self.locked_values.get('check_edt_edp', False)
        self.check_makeup.active = self.locked_values.get('check_makeup', False)
        self.check_pt.active = self.locked_values.get('check_pt', False)
        self.check_es.active = self.locked_values.get('check_es', False)
        self.check_it.active = self.locked_values.get('check_it', False)
        self.slider_value.text = self.locked_values.get('slider_text', '1')
        self.check_und.active = self.locked_values.get('check_und', False)
        self.check_ml.active = self.locked_values.get('check_ml', False)
        self.check_gr.active = self.locked_values.get('check_gr', False)

    def open_dropdown(self, instance):
        if self.dropdown.parent:
            self.dropdown.parent.remove_widget(self.dropdown)
        self.dropdown.open(instance)

    def check_focus(self, dt):
        if self.lock_mode and not self.ean_sku_id.focus:
            self.ean_sku_id.focus = True
        self.highlight_focus()

    def highlight_focus(self):
        if self.ean_sku_id.focus:
            self.ean_sku_id.background_color = (1, 0, 0, 1)  # Rojo si tiene el foco
        else:
            self.ean_sku_id.background_color = (0, 1, 0, 1)  # Verde si no tiene el foco

    def on_request_close(self, *args, **kwargs):
        self.show_exit_confirmation()
        return True  # Evitar el cierre automático

    def show_exit_confirmation(self):
        content = BoxLayout(orientation='vertical', padding=10, spacing=10)
        content.add_widget(Label(text='¿Está seguro de que desea cerrar la aplicación?'))
        button_layout = BoxLayout(size_hint=(1, 0.2))
        yes_button = Button(text='Sí')
        yes_button.bind(on_press=self.confirm_exit)
        no_button = Button(text='No')
        no_button.bind(on_press=lambda x: self.exit_confirmation_popup.dismiss())
        button_layout.add_widget(yes_button)
        button_layout.add_widget(no_button)
        content.add_widget(button_layout)
        self.exit_confirmation_popup = Popup(title='Confirmación de salida',
                                             content=content,
                                             size_hint=(0.6, 0.4))
        self.exit_confirmation_popup.open()

    def confirm_exit(self, instance):
        self.exit_confirmation_popup.dismiss()
        App.get_running_app().stop()

    def on_reg_db_press(self, instance):
        self.reg_db_start_time = datetime.now()
        Clock.schedule_once(self.reg_db_ready, 3)

    def reg_db_ready(self, dt):
        self.status_bar.text = 'Estado: Listo para Importación Masiva'

    def on_reg_db_release(self, instance):
        if (datetime.now() - self.reg_db_start_time).total_seconds() >= 3:
            self.importacion_revs_masiva()
        else:
            self.show_add_to_db_popup()

    def importacion_revs_masiva(self):
        self.file_chooser = CustomFileChooser(filters=['*.xlsx'])  # Configuración inicial para filtrar archivos Excel
        self.file_chooser_layout = BoxLayout(orientation='vertical', spacing=10, padding=10)

        # Campo de texto para filtrar archivos
        self.file_filter_input = TextInput(hint_text='Filtrar archivos...', multiline=False, size_hint=(1, 0.1))
        self.file_filter_input.bind(text=self.filter_files)  # Permitir al usuario filtrar archivos dinámicamente

        # Botón para alternar entre vista de iconos y lista
        self.toggle_view_button = ToggleButton(text='Vista: Iconos', size_hint=(1, 0.1))
        self.toggle_view_button.bind(on_press=self.toggle_file_chooser_view)  # Alternar entre vistas según preferencia

        # ScrollView para la barra de desplazamiento
        self.file_scroll_view = ScrollView(size_hint=(1, 0.8))
        self.file_scroll_view.add_widget(self.file_chooser)

        self.file_chooser_layout.add_widget(self.file_filter_input)
        self.file_chooser_layout.add_widget(self.toggle_view_button)
        self.file_chooser_layout.add_widget(self.file_scroll_view)

        self.file_chooser_popup = Popup(title='Seleccionar archivo .xlsx',
                                        content=self.file_chooser_layout,
                                        size_hint=(0.8, 0.8))
        self.file_chooser.bind(on_submit=self.on_file_selected)  # Vincular evento para manejar selección de archivo
        self.file_chooser_popup.open()

    def filter_files(self, instance, text):
        self.file_chooser.filters = [f'*{text}*'] if text else ['*.xlsx']

    def toggle_file_chooser_view(self, instance):
        # Guardar la ubicación actual antes de cambiar el modo
        current_path = self.file_chooser.path

        # Cambiar entre FileChooserIconView y FileChooserListView
        if isinstance(self.file_chooser, FileChooserIconView):
            self.file_scroll_view.remove_widget(self.file_chooser)
            self.file_chooser = FileChooserListView(filters=['*.xlsx'])
        else:
            self.file_scroll_view.remove_widget(self.file_chooser)
            self.file_chooser = FileChooserIconView(filters=['*.xlsx'])

        # Restaurar la ubicación actual
        self.file_chooser.path = current_path
        self.file_chooser.bind(on_submit=self.on_file_selected)
        self.file_scroll_view.add_widget(self.file_chooser)

        # Actualizar el texto del botón
        self.toggle_view_button.text = 'Vista: Iconos' if isinstance(self.file_chooser, FileChooserListView) else 'Vista: Lista'

    def verify_products_in_db(self, file_path):
        """
        Verifica si los productos del archivo a importar existen en la base de datos.
        Si no existen, solicita confirmación para registrarlos.
        """
        wb = load_workbook(file_path)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        missing_products = []

        for row in rows:
            sku, titulo, eans = row[:3]
            self.cursor.execute('SELECT sku FROM productos WHERE sku = ?', (sku,))
            if not self.cursor.fetchone():
                missing_products.append((sku, titulo, eans))

        if missing_products:
            self.show_missing_products_popup(missing_products, file_path)
        else:
            self.show_import_confirmation(file_path)

    def show_missing_products_popup(self, missing_products, file_path):
        """
        Muestra un popup con los productos que no existen en la base de datos.
        Permite al usuario decidir si desea registrarlos o continuar sin registrarlos.
        """
        content = BoxLayout(orientation='vertical', spacing=10, padding=10)
        scroll_view = ScrollView(size_hint=(1, 0.8))
        # Usar GridLayout para mostrar cada producto en una línea separada y permitir scroll correcto
        from kivy.uix.gridlayout import GridLayout
        products_layout = GridLayout(cols=1, size_hint_y=None, spacing=5, padding=[10, 0, 10, 0])
        products_layout.bind(minimum_height=products_layout.setter('height'))

        for sku, titulo, eans in missing_products:
            label = Label(
                text=f"SKU: {sku}\nTítulo: {titulo}\nEANs: {eans}",
                size_hint_y=None,
                height=70,
                halign='left',
                valign='top',
                text_size=(500, None),
                color=(1, 1, 1, 1)
            )
            label.bind(size=lambda instance, value: setattr(instance, 'text_size', (instance.width, None)))
            products_layout.add_widget(label)

        scroll_view.add_widget(products_layout)
        content.add_widget(scroll_view)

        button_layout = BoxLayout(size_hint=(1, 0.2), spacing=10)
        register_button = Button(text='Registrar en DB')
        register_button.bind(on_press=lambda x: self.register_missing_products(missing_products, file_path))
        continue_button = Button(text='Continuar sin registrar')
        continue_button.bind(on_press=lambda x: self.show_import_confirmation(file_path))
        button_layout.add_widget(register_button)
        button_layout.add_widget(continue_button)

        content.add_widget(button_layout)

        self.missing_products_popup = Popup(
            title='Productos no encontrados en DB',
            content=content,
            size_hint=(0.8, 0.8)
        )
        self.missing_products_popup.open()

    def register_missing_products(self, missing_products, file_path):
        """
        Registra los productos faltantes en la base de datos y continúa con la importación masiva.
        """
        self.missing_products_popup.dismiss()
        self.show_progress_popup('Registrando productos en DB...')

        try:
            for sku, titulo, eans in missing_products:
                self.cursor.execute('INSERT INTO productos (sku, titulo, eans) VALUES (?, ?, ?)', (sku, titulo, eans))
            self.conn.commit()
            self.progress_popup.dismiss()
            self.status_bar.text = f'{len(missing_products)} productos registrados en DB correctamente.'
            self.show_import_confirmation(file_path)
        except sqlite3.OperationalError as e:
            self.progress_popup.dismiss()
            if "database is locked" in str(e):
                msg = "La base de datos está en uso por otro proceso. Por favor, cierre cualquier programa que esté usando 'db.db' y vuelva a intentarlo."
                print(msg)
                self.show_warning_popup(msg)
            else:
                self.show_warning_popup(f'Error al registrar productos en DB: {str(e)}')

    def on_file_selected(self, instance, selection, *args):
        if selection:
            self.file_chooser_popup.dismiss()
            self.verify_products_in_db(selection[0])

    def show_import_confirmation(self, file_path):
        self.import_file_path = file_path
        try:
            # Leer el archivo para obtener un resumen
            wb = load_workbook(file_path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total_rows = len(rows)

            # Resumen de características seleccionadas
            tipo = self.selected_tipo if hasattr(self, 'selected_tipo') else 'ZZ' if self.check_zz.active else 'LOTE' if self.check_lote.active else 'Set & Pack' if self.check_set_pack.active else 'Consumo' if self.check_consumo.active else 'EDT & EDP' if self.check_edt_edp.active else 'MakeUP' if self.check_makeup.active else ''
            tiene_pt = 'Tiene PT' if self.check_pt.active else 'No Tiene PT - TRADUZIDO'
            tiene_es = 'Tiene ES' if self.check_es.active else 'No Tiene ES - TRADUCIDO'
            tiene_it = 'Tiene IT' if self.check_it.active else 'No Tiene IT - TRADOTTO'
            cantidad_neta = self.slider_value.text
            unidad = 'UND' if self.check_und.active else 'ML' if self.check_ml.active else 'GR' if self.check_gr.active else ''

            # Crear el contenido del popup
            content = BoxLayout(orientation='vertical', padding=10, spacing=10)
            scroll_view = ScrollView(size_hint=(1, 0.6))
            summary_layout = BoxLayout(orientation='vertical', size_hint_y=None, spacing=5, padding=5)
            summary_layout.bind(minimum_height=summary_layout.setter('height'))

            # Agregar información del archivo y características seleccionadas
            summary_layout.add_widget(Label(text=f'Archivo seleccionado:\n{file_path}', size_hint_y=None, height=60, halign='left', valign='middle', text_size=(500, None)))
            summary_layout.add_widget(Label(text=f'Total de productos a importar: {total_rows}', size_hint_y=None, height=40, halign='left', valign='middle', text_size=(500, None)))
            summary_layout.add_widget(Label(text='Características seleccionadas:', size_hint_y=None, height=30, bold=True))
            summary_layout.add_widget(Label(text=f'- Tipo: {tipo}', size_hint_y=None, height=30))
            summary_layout.add_widget(Label(text=f'- PT: {tiene_pt}', size_hint_y=None, height=30))
            summary_layout.add_widget(Label(text=f'- ES: {tiene_es}', size_hint_y=None, height=30))
            summary_layout.add_widget(Label(text=f'- IT: {tiene_it}', size_hint_y=None, height=30))
            summary_layout.add_widget(Label(text=f'- Cantidad Neta: {cantidad_neta} {unidad}', size_hint_y=None, height=30))

            scroll_view.add_widget(summary_layout)
            content.add_widget(scroll_view)

            # Botones para confirmar con diferentes estados
            button_layout = BoxLayout(size_hint=(1, 0.2), spacing=10)
            solo_revision_button = Button(text='Solo Revisión')
            solo_revision_button.bind(on_press=lambda x: self.start_mass_import('Solo Revisión'))
            revisado_traducido_button = Button(text='Revisado y Traducido')
            revisado_traducido_button.bind(on_press=lambda x: self.start_mass_import('Revisado y Traducido'))
            button_layout.add_widget(solo_revision_button)
            button_layout.add_widget(revisado_traducido_button)

            content.add_widget(button_layout)

            self.import_confirmation_popup = Popup(title='Confirmación de Importación',
                                                    content=content,
                                                    size_hint=(0.8, 0.8))
            self.import_confirmation_popup.open()
        except Exception as e:
            self.show_warning_popup(f'Error al leer el archivo: {str(e)}')

    def start_mass_import(self, estado):
        self.import_confirmation_popup.dismiss()
        self.show_progress_overlay()
        Clock.schedule_once(lambda dt: self.perform_mass_import(estado), 0.1)

    def show_progress_overlay(self):
        self.progress_overlay = Popup(title='Importando...',
                                       content=ProgressBar(max=100),
                                       size_hint=(0.6, 0.2),
                                       auto_dismiss=False)
        self.progress_overlay.content.value = 0
        self.progress_overlay.open()

    def perform_mass_import(self, estado):
        from openpyxl import load_workbook

        try:
            wb = load_workbook(self.import_file_path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            total_rows = len(rows)
            fecha = datetime.now().strftime('%d-%m-%Y')
            archivo = f'REVs/REV-{fecha}.xlsx'

            if not os.path.exists('REVs'):
                os.makedirs('REVs')

            if os.path.exists(archivo):
                rev_wb = load_workbook(archivo)
                rev_ws = rev_wb.active
            else:
                rev_wb = Workbook()
                rev_ws = rev_wb.active
                rev_ws.append(['EAN/SKU/ID', 'MARCA/TITULO', 'Tipo', 'Tiene PT', 'Tiene ES', 'Tiene IT', 'Cantidad Neta', 'UND/ML/GR', 'Composición de Lote', 'Estado'])

            imported_count = 0

            for i, row in enumerate(rows):
                try:
                    sku, titulo, eans = row
                    tipo = self.selected_tipo if hasattr(self, 'selected_tipo') else 'ZZ' if self.check_zz.active else 'LOTE' if self.check_lote.active else 'Set & Pack' if self.check_set_pack.active else 'Consumo' if self.check_consumo.active else 'EDT & EDP' if self.check_edt_edp.active else 'MakeUP' if self.check_makeup.active else ''
                    tiene_pt = 'Tiene PT' if self.check_pt.active else 'No Tiene PT - TRADUZIDO'
                    tiene_es = 'Tiene ES' if self.check_es.active else 'No Tiene ES - TRADUCIDO'
                    tiene_it = 'Tiene IT' if self.check_it.active else 'No Tiene IT - TRADOTTO'
                    cantidad_neta = self.slider_value.text
                    unidad = 'UND' if self.check_und.active else 'ML' if self.check_ml.active else 'GR' if self.check_gr.active else ''
                    composicion_lote = self.lote_composition if self.check_lote.active or self.check_set_pack.active else ''

                    rev_ws.append([sku, titulo, tipo, tiene_pt, tiene_es, tiene_it, cantidad_neta, unidad, composicion_lote, estado])
                    imported_count += 1
                    if hasattr(self, 'progress_overlay'):  # Verificar si el popup existe
                        self.progress_overlay.content.value = int((i + 1) / total_rows * 100)
                    self.root.do_layout()
                except Exception as e:
                    self.show_warning_popup(f'Error al importar el producto en la fila {i + 2}: {str(e)}')

            rev_wb.save(archivo)
            if hasattr(self, 'progress_overlay'):  # Verificar si el popup existe
                self.progress_overlay.dismiss()
            self.status_bar.text = f'Importación Masiva Completada: {imported_count} productos'
        except Exception as e:
            if hasattr(self, 'progress_overlay'):  # Verificar si el popup existe
                self.progress_overlay.dismiss()
            self.show_warning_popup(f'Error durante la importación: {str(e)}')

    def on_marca_titulo_enter(self, instance):
        """
        Maneja el evento de presionar Enter en el campo de texto "Marca/Titulo".
        Realiza una búsqueda en la base de datos utilizando las palabras clave ingresadas.
        """
        keywords = self.marca_titulo.text.strip().split()
        if not keywords:
            self.show_warning_popup('El campo Marca/Titulo\nno puede estar vacío.')
            return

        # Mostrar barra de progreso mientras se realiza la búsqueda
        self.show_progress_popup('Buscando productos...')

        try:
            # Construir la consulta SQL para buscar coincidencias en la columna "titulo"
            self.cursor.execute('SELECT sku, titulo FROM productos WHERE ' + ' AND '.join(["titulo LIKE ?" for _ in keywords]), [f'%{kw}%' for kw in keywords])
            results = self.cursor.fetchall()
        except sqlite3.OperationalError as e:
            self.progress_popup.dismiss()
            if "database is locked" in str(e):
                msg = "La base de datos está en uso por otro proceso. Por favor, cierre cualquier programa que esté usando 'db.db' y vuelva a intentarlo."
                print(msg)
                self.show_warning_popup(msg)
                return
            else:
                raise

        # Cerrar la barra de progreso
        self.progress_popup.dismiss()

        if results:
            self.show_results_popup(results)
        else:
            self.show_warning_popup('No se encontraron productos que coincidan con las palabras clave.')

    def show_progress_popup(self, message):
        """
        Muestra un popup con una barra de progreso y un mensaje.

        Argumentos:
        - message: Mensaje a mostrar en el popup.
        """
        content = BoxLayout(orientation='vertical', padding=10, spacing=10)
        content.add_widget(Label(text=message, size_hint=(1, 0.2)))
        progress_bar = ProgressBar(max=100, value=50, size_hint=(1, 0.2))
        content.add_widget(progress_bar)
        self.progress_popup = Popup(title='Procesando',
                                    content=content,
                                    size_hint=(0.6, 0.4),
                                    auto_dismiss=False)
        self.progress_popup.open()

    def show_results_popup(self, results):
        """
        Muestra un popup interactivo con los resultados de la búsqueda, con paginación y exportación.

        Argumentos:
        - results: Lista de tuplas con los resultados de la búsqueda. Cada tupla contiene (sku, titulo).
        """
        self.results_full_list = results
        self.results_popup_index = 0  # Índice de inicio del bloque actual

        def get_current_block():
            start = self.results_popup_index
            end = min(start + self.RESULTS_BLOCK_SIZE, len(self.results_full_list))
            return self.results_full_list[start:end]

        def update_results_layout():
            results_layout.clear_widgets()
            for sku, titulo in get_current_block():
                result_button = Button(text=f"{sku} - {titulo}", size_hint_y=None, height=44)
                result_button.bind(on_release=lambda btn, s=sku, t=titulo: self.select_result(s, t))
                results_layout.add_widget(result_button)
            # Deshabilitar "Cargar más" si no hay más resultados
            if self.results_popup_index + self.RESULTS_BLOCK_SIZE >= len(self.results_full_list):
                load_more_button.disabled = True
            else:
                load_more_button.disabled = False

        def on_load_more(instance):
            self.results_popup_index += self.RESULTS_BLOCK_SIZE
            update_results_layout()

        def on_export(instance):
            self.export_results_to_xlsx(self.results_full_list)

        # --- Layout del popup ---
        content = BoxLayout(orientation='vertical', spacing=10, padding=10)
        scroll_view = ScrollView(size_hint=(1, 0.7))
        results_layout = BoxLayout(orientation='vertical', size_hint_y=None)
        results_layout.bind(minimum_height=results_layout.setter('height'))
        scroll_view.add_widget(results_layout)
        content.add_widget(scroll_view)

        # Botones inferiores
        buttons_layout = BoxLayout(size_hint=(1, 0.15), spacing=10)
        load_more_button = Button(text='Cargar más')
        load_more_button.bind(on_press=on_load_more)
        export_button = Button(text='EXPORT')
        export_button.bind(on_press=on_export)
        close_button = Button(text='Cerrar')
        close_button.bind(on_press=lambda x: self.results_popup.dismiss())
        buttons_layout.add_widget(load_more_button)
        buttons_layout.add_widget(export_button)
        buttons_layout.add_widget(close_button)
        content.add_widget(buttons_layout)

        num_results = len(results)
        self.results_popup = Popup(
            title=f'Resultados de la búsqueda ({num_results}) Productos Encontrados MATCHs',
            content=content,
            size_hint=(0.8, 0.8)
        )
        update_results_layout()
        self.results_popup.open()

    def export_results_to_xlsx(self, results):
        """
        Exporta la lista de resultados a un archivo xlsx, permitiendo al usuario elegir la ubicación y nombre.
        Incluye las columnas SKU, TITULO y EANs actuales del producto.
        """
        from kivy.uix.filechooser import FileChooserIconView
        from kivy.uix.textinput import TextInput

        def on_export_confirm(instance):
            export_path = file_chooser.path
            filename = filename_input.text.strip()
            if not filename.lower().endswith('.xlsx'):
                filename += '.xlsx'
            full_path = os.path.join(export_path, filename)
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.append(['SKU', 'TITULO', 'EANs'])
                # Obtener los EANs de la base de datos para cada SKU
                for sku, titulo in results:
                    self.cursor.execute('SELECT eans FROM productos WHERE sku = ?', (sku,))
                    row = self.cursor.fetchone()
                    eans = row[0] if row and row[0] else ''
                    ws.append([sku, titulo, eans])
                wb.save(full_path)
                export_popup.dismiss()
                self.show_info_popup('Exportación exitosa', f'Archivo exportado:\n{full_path}')
            except Exception as e:
                self.show_warning_popup(f'Error al exportar: {str(e)}')

        file_chooser = FileChooserIconView(filters=['*.xlsx'], size_hint=(1, 0.7))
        filename_input = TextInput(hint_text='Nombre del archivo (sin extensión)', multiline=False, size_hint=(1, 0.1))
        export_btn = Button(text='Exportar', size_hint=(1, 0.1))
        export_btn.bind(on_press=on_export_confirm)
        cancel_btn = Button(text='Cancelar', size_hint=(1, 0.1))
        cancel_btn.bind(on_press=lambda x: export_popup.dismiss())

        layout = BoxLayout(orientation='vertical', spacing=10, padding=10)
        layout.add_widget(file_chooser)
        layout.add_widget(filename_input)
        layout.add_widget(export_btn)
        layout.add_widget(cancel_btn)

        export_popup = Popup(title='Exportar resultados a XLSX', content=layout, size_hint=(0.8, 0.8))
        export_popup.open()

    def on_f4k3_press(self, instance):
        """
        Al hacer click en F4K3, registra una revisión predefinida en el archivo REVs del día actual.
        Los valores son de ejemplo y pueden ser modificados posteriormente.
        Permite duplicar la entrada tantas veces como se pulse el botón.
        """
        from openpyxl import load_workbook, Workbook
        from datetime import datetime
        import os

        # Valores predefinidos de ejemplo (sin Composición de Lote)
        fake_row = [
            "FAKE",                         # EAN/SKU/ID
            "Marca Falsa - Producto Fake",  # MARCA/TITULO
            "VARIOS",                       # Tipo
            "Tiene PT",                     # Tiene PT
            "Tiene ES",                     # Tiene ES
            "Tiene IT",                     # Tiene IT
            "1",                            # Cantidad Neta
            "UND",                          # UND/ML/GR
            "",                             # Composición de Lote vacío
            "Solo Revisión"                 # Estado
        ]

        fecha = datetime.now().strftime('%d-%m-%Y')
        archivo = f'REVs/REV-{fecha}.xlsx'

        if not os.path.exists('REVs'):
            os.makedirs('REVs')

        try:
            if os.path.exists(archivo):
                wb = load_workbook(archivo)
                ws = wb.active
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(['EAN/SKU/ID', 'MARCA/TITULO', 'Tipo', 'Tiene PT', 'Tiene ES', 'Tiene IT', 'Cantidad Neta', 'UND/ML/GR', 'Composición de Lote', 'Estado'])
            ws.append(fake_row)
            wb.save(archivo)
            self.status_bar.text = 'Estado: Registro F4K3 añadido'
        except Exception as e:
            self.show_warning_popup(f'Error al registrar F4K3: {str(e)}')

if __name__ == '__main__':
    try:
        ContadorApp().run()
    except KeyboardInterrupt:
        print("Aplicación cerrada por el usuario.")
    except Exception as e:
        import traceback
        print("Ocurrió un error inesperado:")
        traceback.print_exc()
        input("Presione Enter para cerrar...")

